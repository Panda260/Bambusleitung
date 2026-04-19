"""
exporter.py – Excel-Export mit openpyxl
"""
import io
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def export_to_excel(results: list) -> bytes:
    """
    Erstellt eine formatierte Excel-Datei aus der History-Liste.
    Gibt bytes zurück (für Flask send_file).
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Speedtest History"

    # ---- Header-Style ----
    header_font = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
    header_fill = PatternFill(fill_type="solid", fgColor="1A1A2E")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC"),
    )
    alt_fill = PatternFill(fill_type="solid", fgColor="F0F4FF")

    headers = [
        ("ID", 6),
        ("Zeitstempel", 22),
        ("Typ", 10),
        ("Ziel-IP", 18),
        ("Port", 8),
        ("Download (Mbit/s)", 18),
        ("Upload (Mbit/s)", 16),
        ("Jitter (ms)", 12),
        ("Paketverlust (%)", 16),
        ("Dauer (s)", 10),
        ("Retransmits", 13),
        ("Status", 10),
        ("Fehler", 35),
    ]

    # ---- Titelzeile ----
    ws.merge_cells("A1:M1")
    title_cell = ws["A1"]
    title_cell.value = f"🎋 Bambusleitung – Speedtest Export  |  Erstellt: {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')}"
    title_cell.font = Font(bold=True, size=13, color="FFFFFF", name="Calibri")
    title_cell.fill = PatternFill(fill_type="solid", fgColor="0F3460")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    # ---- Header-Zeile ----
    for col_idx, (header, width) in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[2].height = 32

    # ---- Daten-Zeilen ----
    for row_idx, entry in enumerate(results, start=3):
        is_alt = (row_idx % 2 == 0)
        row_data = [
            entry.get("id"),
            entry.get("timestamp"),
            entry.get("run_type", "").capitalize(),
            entry.get("target_ip"),
            entry.get("target_port"),
            entry.get("download_mbps"),
            entry.get("upload_mbps"),
            entry.get("jitter_ms"),
            entry.get("packet_loss_pct"),
            entry.get("duration_s"),
            entry.get("retransmits"),
            entry.get("status", "").capitalize(),
            entry.get("error_msg") or "",
        ]
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center")
            if is_alt:
                cell.fill = alt_fill
            # Zahlen-Format für Mbps
            if col_idx in (6, 7) and isinstance(value, float):
                cell.number_format = "0.00"
            # Status-Farbe
            if col_idx == 12:
                if str(value).lower() == "success":
                    cell.font = Font(color="1D7A3A", bold=True, name="Calibri")
                elif str(value).lower() == "error":
                    cell.font = Font(color="C0392B", bold=True, name="Calibri")
        ws.row_dimensions[row_idx].height = 20

    # ---- Freeze header ----
    ws.freeze_panes = "A3"

    # ---- Auto-Filter ----
    ws.auto_filter.ref = f"A2:M{max(2, len(results) + 2)}"

    # ---- In Bytes serialisieren ----
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.read()
